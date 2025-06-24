from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.utils import timezone
from django.db.models import Q
from django.middleware.csrf import get_token
from django.views.decorators.csrf import ensure_csrf_cookie
from django.utils.decorators import method_decorator
from datetime import date, timedelta, datetime
import calendar
import pytz
from openpyxl import Workbook
from django.http import HttpResponse
from .models import Employee, TimeRecord, MonthlyReport
from .serializers import EmployeeSerializer, TimeRecordSerializer, MonthlyReportSerializer

@method_decorator(ensure_csrf_cookie, name='dispatch')
class AuthViewSet(viewsets.ViewSet):
    permission_classes = []
    
    @action(detail=False, methods=['get'])
    def csrf(self, request):
        """Get CSRF token"""
        return Response({'csrfToken': get_token(request)})
    
    @action(detail=False, methods=['post'])
    def login(self, request):
        username = request.data.get('username')
        password = request.data.get('password')
        
        user = authenticate(username=username, password=password)
        if user:
            login(request, user)
            try:
                employee = Employee.objects.get(user=user)
                return Response({
                    'success': True,
                    'employee': EmployeeSerializer(employee).data
                })
            except Employee.DoesNotExist:
                return Response({'success': False, 'message': 'Employee profile not found'}, 
                              status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({'success': False, 'message': 'Invalid credentials'}, 
                          status=status.HTTP_401_UNAUTHORIZED)
    
    @action(detail=False, methods=['post'])
    def logout(self, request):
        logout(request)
        return Response({'success': True, 'message': 'Logged out successfully'})
    
    @action(detail=False, methods=['get'])
    def status(self, request):
        """Check authentication status without requiring authentication"""
        if request.user.is_authenticated:
            try:
                employee = Employee.objects.get(user=request.user)
                return Response({
                    'authenticated': True,
                    'employee': EmployeeSerializer(employee).data
                })
            except Employee.DoesNotExist:
                return Response({
                    'authenticated': False,
                    'message': 'Employee profile not found'
                })
        else:
            return Response({'authenticated': False})
        return Response({'success': True, 'message': 'Logged out successfully'})

class EmployeeViewSet(viewsets.ModelViewSet):
    queryset = Employee.objects.all()
    serializer_class = EmployeeSerializer
    
    @action(detail=False, methods=['get'])
    def current(self, request):
        if request.user.is_authenticated:
            try:
                employee = Employee.objects.get(user=request.user)
                return Response(EmployeeSerializer(employee).data)
            except Employee.DoesNotExist:
                return Response({'message': 'Employee profile not found'}, 
                              status=status.HTTP_404_NOT_FOUND)
        return Response({'message': 'Not authenticated'}, status=status.HTTP_401_UNAUTHORIZED)

class TimeRecordViewSet(viewsets.ModelViewSet):
    queryset = TimeRecord.objects.all()
    serializer_class = TimeRecordSerializer
    
    def get_queryset(self):
        if self.request.user.is_authenticated:
            try:
                employee = Employee.objects.get(user=self.request.user)
                return TimeRecord.objects.filter(employee=employee)
            except Employee.DoesNotExist:
                return TimeRecord.objects.none()
        return TimeRecord.objects.none()
    
    @action(detail=False, methods=['post'])
    def checkin_checkout(self, request):
        """Smart check-in/checkout logic with forgotten checkout handling"""
        try:
            employee = Employee.objects.get(user=request.user)
        except Employee.DoesNotExist:
            return Response({'message': 'Employee profile not found'}, 
                          status=status.HTTP_404_NOT_FOUND)
        
        # Prevent admin users from checking in/out
        if employee.employee_id == 'ADMIN001' or employee.department == 'HR' and employee.position == 'System Administrator':
            return Response({
                'success': False,
                'message': 'Admin users cannot check in/out. Use the Admin Dashboard to manage employee data.'
            }, status=status.HTTP_403_FORBIDDEN)
        
        today = timezone.now().date()
        current_time = timezone.now()
        
        # Check if there's a record for today
        today_record, created = TimeRecord.objects.get_or_create(
            employee=employee,
            date=today,
            defaults={'status': 'CHECKED_OUT'}
        )
        
        # Check if employee forgot to checkout yesterday
        yesterday = today - timedelta(days=1)
        try:
            yesterday_record = TimeRecord.objects.get(employee=employee, date=yesterday)
            if yesterday_record.status == 'CHECKED_IN':
                # Mark yesterday as forgot checkout
                yesterday_record.status = 'FORGOT_CHECKOUT'
                yesterday_record.forgot_checkout = True
                yesterday_record.save()
        except TimeRecord.DoesNotExist:
            pass
        
        # Current action logic
        if today_record.status == 'CHECKED_OUT':
            # Check in
            today_record.check_in_time = current_time
            today_record.status = 'CHECKED_IN'
            today_record.save()
            
            # Format time for Vietnam timezone display
            vietnam_tz = pytz.timezone('Asia/Ho_Chi_Minh')
            vietnam_time = current_time.astimezone(vietnam_tz)
            formatted_time = vietnam_time.strftime("%H:%M:%S")
            
            return Response({
                'success': True,
                'action': 'checked_in',
                'message': f'Checked in at {formatted_time}',
                'record': TimeRecordSerializer(today_record).data
            })
        elif today_record.status == 'CHECKED_IN':
            # Check out
            today_record.check_out_time = current_time
            today_record.status = 'CHECKED_OUT'
            today_record.calculate_working_hours()
            today_record.save()
            
            # Format time for Vietnam timezone display
            vietnam_tz = pytz.timezone('Asia/Ho_Chi_Minh')
            vietnam_time = current_time.astimezone(vietnam_tz)
            formatted_time = vietnam_time.strftime("%H:%M:%S")
            
            return Response({
                'success': True,
                'action': 'checked_out',
                'message': f'Checked out at {formatted_time} - Worked {today_record.working_hours} hours',
                'record': TimeRecordSerializer(today_record).data
            })
        else:
            return Response({
                'success': False,
                'message': 'Invalid status'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def current_status(self, request):
        """Get current check-in status"""
        try:
            employee = Employee.objects.get(user=request.user)
        except Employee.DoesNotExist:
            return Response({'message': 'Employee profile not found'}, 
                          status=status.HTTP_404_NOT_FOUND)
        
        # Return special status for admin users
        if employee.employee_id == 'ADMIN001' or employee.department == 'HR' and employee.position == 'System Administrator':
            return Response({
                'status': 'ADMIN',
                'record': None,
                'is_admin': True,
                'message': 'Admin users do not track time. Access Admin Dashboard for system management.'
            })
        
        today = timezone.now().date()
        try:
            today_record = TimeRecord.objects.get(employee=employee, date=today)
            return Response({
                'status': today_record.status,
                'record': TimeRecordSerializer(today_record).data,
                'is_admin': False
            })
        except TimeRecord.DoesNotExist:
            return Response({
                'status': 'CHECKED_OUT',
                'record': None,
                'is_admin': False
            })
    
    @action(detail=False, methods=['get'])
    def monthly_records(self, request):
        """Get monthly time records"""
        try:
            employee = Employee.objects.get(user=request.user)
        except Employee.DoesNotExist:
            return Response({'message': 'Employee profile not found'}, 
                          status=status.HTTP_404_NOT_FOUND)
        
        year = int(request.query_params.get('year', timezone.now().year))
        month = int(request.query_params.get('month', timezone.now().month))
        
        records = TimeRecord.objects.filter(
            employee=employee,
            date__year=year,
            date__month=month
        ).order_by('-date')
        
        return Response(TimeRecordSerializer(records, many=True).data)

class AdminViewSet(viewsets.ViewSet):
    """Admin-only endpoints for system-wide data management"""
    
    def _is_admin(self, user):
        """Check if user has admin privileges"""
        return user.is_authenticated and (user.is_superuser or user.is_staff)
    
    @action(detail=False, methods=['get'])
    def all_employees(self, request):
        """Get all employees data - Admin only"""
        if not self._is_admin(request.user):
            return Response({'message': 'Admin access required'}, 
                          status=status.HTTP_403_FORBIDDEN)
        
        employees = Employee.objects.filter(is_active=True)
        return Response(EmployeeSerializer(employees, many=True).data)
    
    @action(detail=False, methods=['get'])
    def system_stats(self, request):
        """Get system-wide statistics - Admin only"""
        if not self._is_admin(request.user):
            return Response({'message': 'Admin access required'}, 
                          status=status.HTTP_403_FORBIDDEN)
        
        today = timezone.now().date()
        
        # Basic stats
        total_employees = Employee.objects.filter(is_active=True).count()
        checked_in_today = TimeRecord.objects.filter(
            date=today,
            status='CHECKED_IN'
        ).count()
        
        # Monthly stats
        current_month_records = TimeRecord.objects.filter(
            date__year=today.year,
            date__month=today.month
        )
        
        total_working_hours_month = sum([r.working_hours for r in current_month_records])
        forgotten_checkouts_month = current_month_records.filter(forgot_checkout=True).count()
        
        return Response({
            'total_employees': total_employees,
            'checked_in_today': checked_in_today,
            'checked_out_today': total_employees - checked_in_today,
            'total_working_hours_this_month': round(total_working_hours_month, 2),
            'forgotten_checkouts_this_month': forgotten_checkouts_month,
            'month_year': f"{today.strftime('%B')} {today.year}"
        })
    
    @action(detail=False, methods=['get'])
    def all_employees_records(self, request):
        """Get all employees' time records for a specific month - Admin only"""
        if not self._is_admin(request.user):
            return Response({'message': 'Admin access required'}, 
                          status=status.HTTP_403_FORBIDDEN)
        
        year = int(request.query_params.get('year', timezone.now().year))
        month = int(request.query_params.get('month', timezone.now().month))
        
        employees_data = []
        employees = Employee.objects.filter(is_active=True)
        
        for employee in employees:
            records = TimeRecord.objects.filter(
                employee=employee,
                date__year=year,
                date__month=month
            )
            
            total_working_days = records.filter(check_in_time__isnull=False).count()
            total_working_hours = sum([r.working_hours for r in records])
            days_forgot_checkout = records.filter(forgot_checkout=True).count()
            
            # Calculate days off (total days in month - working days)
            days_in_month = calendar.monthrange(year, month)[1]
            days_off = days_in_month - total_working_days
            
            employees_data.append({
                'employee': EmployeeSerializer(employee).data,
                'stats': {
                    'total_working_days': total_working_days,
                    'total_working_hours': round(total_working_hours, 2),
                    'days_forgot_checkout': days_forgot_checkout,
                    'days_off': days_off,
                    'records': TimeRecordSerializer(records, many=True).data
                }
            })
        
        return Response({
            'year': year,
            'month': month,
            'employees_data': employees_data
        })
    
    @action(detail=False, methods=['get'])
    def comprehensive_excel(self, request):
        """Generate comprehensive Excel report for all employees - Admin only"""
        if not self._is_admin(request.user):
            return Response({'message': 'Admin access required'}, 
                          status=status.HTTP_403_FORBIDDEN)
        
        year = int(request.query_params.get('year', timezone.now().year))
        month = int(request.query_params.get('month', timezone.now().month))
        
        # Create workbook with multiple sheets
        wb = Workbook()
        
        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Summary headers
        summary_headers = [
            'Employee ID', 'Full Name', 'Department', 'Position',
            'Total Working Days', 'Total Working Hours', 'Days Forgot Checkout',
            'Days Off', 'Average Hours/Day', 'Status'
        ]
        ws_summary.append(summary_headers)
        
        # Detailed Sheet
        ws_detailed = wb.create_sheet("Detailed Records")
        detailed_headers = [
            'Employee ID', 'Full Name', 'Department', 'Date',
            'Check In', 'Check Out', 'Working Hours', 'Status', 'Notes'
        ]
        ws_detailed.append(detailed_headers)
        
        # Get all employees and their data
        employees = Employee.objects.filter(is_active=True)
        
        for employee in employees:
            records = TimeRecord.objects.filter(
                employee=employee,
                date__year=year,
                date__month=month
            ).order_by('date')
            
            total_working_days = records.filter(check_in_time__isnull=False).count()
            total_working_hours = sum([r.working_hours for r in records])
            days_forgot_checkout = records.filter(forgot_checkout=True).count()
            
            # Calculate stats
            days_in_month = calendar.monthrange(year, month)[1]
            days_off = days_in_month - total_working_days
            avg_hours_per_day = round(total_working_hours / total_working_days, 2) if total_working_days > 0 else 0
            
            # Employee status
            status_text = "Active"
            if days_forgot_checkout > 5:
                status_text = "Needs Attention"
            elif total_working_hours < 40:
                status_text = "Low Hours"
            
            # Add to summary sheet
            summary_row = [
                employee.employee_id,
                employee.full_name,
                employee.get_department_display(),
                employee.position,
                total_working_days,
                round(total_working_hours, 2),
                days_forgot_checkout,
                days_off,
                avg_hours_per_day,
                status_text
            ]
            ws_summary.append(summary_row)
            
            # Add detailed records
            for record in records:
                detailed_row = [
                    employee.employee_id,
                    employee.full_name,
                    employee.get_department_display(),
                    record.date.strftime('%Y-%m-%d'),
                    record.check_in_time.strftime('%H:%M:%S') if record.check_in_time else 'N/A',
                    record.check_out_time.strftime('%H:%M:%S') if record.check_out_time else 'N/A',
                    record.working_hours,
                    record.get_status_display(),
                    'Forgot checkout' if record.forgot_checkout else ''
                ]
                ws_detailed.append(detailed_row)
        
        # Create response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="admin_comprehensive_report_{month}_{year}.xlsx"'
        wb.save(response)
        return response

class ReportViewSet(viewsets.ViewSet):
    
    @action(detail=False, methods=['get'])
    def monthly_excel(self, request):
        """Generate monthly Excel report for all employees"""
        year = int(request.query_params.get('year', timezone.now().year))
        month = int(request.query_params.get('month', timezone.now().month))
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Report {month}-{year}"
        
        # Headers
        headers = [
            'Employee ID', 'Full Name', 'Department', 'Position',
            'Total Working Days', 'Total Working Hours', 'Days Forgot Checkout',
            'Days Off', 'Notes'
        ]
        ws.append(headers)
        
        # Get all employees
        employees = Employee.objects.filter(is_active=True)
        
        for employee in employees:
            records = TimeRecord.objects.filter(
                employee=employee,
                date__year=year,
                date__month=month
            )
            
            total_working_days = records.filter(check_in_time__isnull=False).count()
            total_working_hours = sum([r.working_hours for r in records])
            days_forgot_checkout = records.filter(forgot_checkout=True).count()
            
            # Calculate days off (total days in month - working days)
            days_in_month = calendar.monthrange(year, month)[1]
            days_off = days_in_month - total_working_days
            
            row = [
                employee.employee_id,
                employee.full_name,
                employee.get_department_display(),
                employee.position,
                total_working_days,
                round(total_working_hours, 2),
                days_forgot_checkout,
                days_off,
                f"Report for {month}/{year}"
            ]
            ws.append(row)
        
        # Create response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="monthly_report_{month}_{year}.xlsx"'
        wb.save(response)
        return response